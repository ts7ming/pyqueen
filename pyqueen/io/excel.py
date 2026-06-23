import pandas as pd
import numpy as np
import os


class Excel:
    def __init__(self, file_path=None):
        self.__ds = None
        if file_path is None:
            file_path = 'pyqueen_excel'
        self.file_path = self.__check_path(file_path=file_path)

    @staticmethod
    def __check_path(file_path):
        if os.path.isabs(file_path) and str(file_path)[-5:] == '.xlsx':
            pass
        elif os.path.isabs(file_path) is False and str(file_path)[-5:] != '.xlsx':
            file_path = os.path.join(os.path.curdir, file_path) + '.xlsx'
        else:
            file_path = os.path.join(os.path.curdir, file_path)
        return file_path

    def read_excel(self, sheet_name=None, file_path=None):
        if file_path is not None:
            file_path = self.__check_path(file_path)
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        return df

    def read_sql(self, sql, data=None, engine='sqlite'):
        from pyqueen import DataSource
        if engine == 'sqlite':
            self.__ds = DataSource(conn_type='sqlite', keep_conn=True)
            if data is None:
                data = pd.read_excel(self.file_path, sheet_name=None)
            for sht_name, df in data.items():
                self.__ds.to_db(df, sht_name)
            df_result = self.__ds.read_sql(sql)
            self.__ds.close_conn()
        elif engine == 'duckdb':
            import duckdb
            with duckdb.connect() as conn:
                for df_name, df in data.items():
                    conn.register(df_name, df)
                df_result = conn.execute(sql).df()
        else:
            raise Exception('wrong engine')
        return df_result

    def to_excel(self, sheet_list, file_path=None, fillna='', fmt=None, font='微软雅黑', font_color='black', font_size=11, column_width=17):
        if file_path is not None:
            file_path = self.__check_path(file_path)
        else:
            file_path = self.file_path

        import xlsxwriter
        if os.path.exists(os.path.dirname(file_path)) is False:
            os.makedirs(os.path.dirname(file_path))
        wb = xlsxwriter.Workbook(file_path)
        fmt_default = wb.add_format()
        fmt_default.set_font_name(font)
        fmt_default.set_font_color(font_color)
        fmt_default.set_font_size(font_size)

        style = {
            'align': 'center',
            'bold': True
        }
        fmt_head = wb.add_format(style)
        fmt_head.set_font_name(font)
        fmt_head.set_font_color(font_color)
        fmt_head.set_font_size(font_size)

        if fmt is None:
            fmt = {}
        for df, sheet_name in sheet_list:
            df = df.fillna(fillna)
            df.replace(-np.inf, fillna, inplace=True)
            df.replace(np.inf, fillna, inplace=True)
            ws = wb._add_sheet(sheet_name)
            head = df.columns.to_list()
            data = df.values.tolist()

            row_num, col_num = df.shape
            for col_name, col_index in zip(head, range(col_num)):
                if col_name in fmt.keys():
                    col_format = None
                    col_format = wb.add_format({'num_format': fmt[col_name]})
                    col_format.set_font_name(font)
                    col_format.set_font_color(font_color)
                    col_format.set_font_size(font_size)
                else:
                    col_format = fmt_default
                ws.set_column(col_index, col_index, column_width)
                ws.write(0, col_index, str(col_name), fmt_head)
                for row_index in range(row_num):
                    value = data[row_index][col_index]
                    ws.write(row_index + 1, col_index, value, col_format)
        wb.close()
        return file_path


    def update_excel(self, file_path=None, script=None):
        import openpyxl
        import re
        from openpyxl.utils import column_index_from_string
        """
        执行Excel更新脚本
        
        Args:
            file_path: Excel文件路径
            script: 操作脚本列表，每个操作是一个字典，支持以下action:
                - v2c: 将指定值填入指定单元格
                - df2c_v: 将DataFrame的值填入Excel（仅数据，不包含表头），从指定单元格开始
                - df2c_hv: 将DataFrame填入Excel（包含表头和数据），从指定单元格开始
                - update: 刷新Excel中所有公式（设置标志让Excel打开时重新计算公式）
                    - sheet (可选): 指定工作表名称，只刷新该工作表的公式
                    - r (可选): 指定单元格或范围，如 "B3" 或 "A3:B6"，只刷新指定范围的公式
        
        Note:
            - openpyxl 无法直接计算公式值，只能存储公式
            - update 操作会设置工作簿属性，使Excel在打开文件时自动重新计算所有公式
            - 如果需要在Python中看到公式计算结果，需要使用其他库如 xlwings 或 comtypes
            - update 支持三种模式：
              1. 全局刷新: {"action": "update"} - 刷新所有工作表的所有公式
              2. 工作表刷新: {"action": "update", "sheet": "Sheet1"} - 只刷新指定工作表
              3. 范围刷新: {"action": "update", "sheet": "Sheet1", "r": "A3:B6"} - 只刷新指定范围
        """

        if file_path is not None:
            file_path = self.__check_path(file_path)
        else:
            file_path = self.file_path
        # 加载工作簿
        try:
            wb = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            # 如果文件不存在，创建新工作簿
            wb = openpyxl.Workbook()
        
        def get_or_create_sheet(wb, sheet_name):
            if sheet_name in wb.sheetnames:
                return wb[sheet_name]
            else:
                return wb.create_sheet(sheet_name)

        def parse_start_cell(start_cell):
            match = re.match(r'([A-Z]+)(\d+)', start_cell)
            if match:
                start_col_letter, start_row = match.groups()
                start_col = column_index_from_string(start_col_letter)
                start_row = int(start_row)
                return start_col, start_row
            return 1, 1 # Default to A1 if parsing fails

        # 执行脚本中的每个操作
        for action_dict in script:
            action = action_dict.get('action')
            
            if action == 'v2c':
                # 将值填入指定单元格
                value = action_dict.get('v')
                sheet_name = action_dict.get('sheet', wb.active.title)
                cell_ref = action_dict.get('c')
                
                ws = get_or_create_sheet(wb, sheet_name)
                
                # 解析单元格引用（如 "B2"）
                if cell_ref:
                    ws[cell_ref] = value
                    
            elif action in ['df2c', 'df2c_hv']:
                # 将DataFrame填入Excel（包含表头和数据）
                df = action_dict.get('df')
                sheet_name = action_dict.get('sheet', wb.active.title)
                start_cell = action_dict.get('c', 'A1')
                
                if df is not None and not df.empty:
                    ws = get_or_create_sheet(wb, sheet_name)
                    start_col, start_row = parse_start_cell(start_cell)
                    
                    # 写入列名（DataFrame的列索引）
                    for col_idx, col_name in enumerate(df.columns):
                        ws.cell(row=start_row, column=start_col + col_idx, value=col_name)
                    
                    # 写入数据
                    for row_idx, row in enumerate(df.values, start=1):
                        for col_idx, value in enumerate(row):
                            ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)
                                
            elif action == 'df2c_v':
                # 将DataFrame填入Excel（仅数据，不包含表头）
                df = action_dict.get('df')
                sheet_name = action_dict.get('sheet', wb.active.title)
                start_cell = action_dict.get('c', 'A1')
                
                if df is not None and not df.empty:
                    ws = get_or_create_sheet(wb, sheet_name)
                    start_col, start_row = parse_start_cell(start_cell)
                    
                    # 仅写入数据（不包含表头）
                    for row_idx, row in enumerate(df.values, start=0):
                        for col_idx, value in enumerate(row):
                            ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)
            
            elif action == 'update':
                # 刷新公式（强制重新计算）
                # 注意：openpyxl 不会自动计算公式值，但可以设置标志让Excel打开时重新计算
                sheet_name = action_dict.get('sheet')
                cell_range = action_dict.get('r')  # 支持 "B3" 或 "A3:B6"
                
                if sheet_name and cell_range:
                    # 刷新指定工作表的指定单元格范围
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        # 解析单元格范围
                        if ':' in cell_range:
                            # 范围格式如 "A3:B6"
                            start_cell, end_cell = cell_range.split(':')
                            for row in ws.iter_rows(min_row=ws[start_cell].row, 
                                                   max_row=ws[end_cell].row,
                                                   min_col=ws[start_cell].column,
                                                   max_col=ws[end_cell].column):
                                for cell in row:
                                    if cell.data_type == 'f':
                                        formula = cell.value
                                        if formula and not str(formula).startswith('='):
                                            cell.value = f"={formula}"
                        else:
                            # 单个单元格如 "B3"
                            cell = ws[cell_range]
                            if cell.data_type == 'f':
                                formula = cell.value
                                if formula and not str(formula).startswith('='):
                                    cell.value = f"={formula}"
                else:
                    # 刷新所有工作表的所有公式
                    target_sheets = [sheet_name] if sheet_name else wb.sheetnames
                    for sht_name in target_sheets:
                        if sht_name in wb.sheetnames:
                            ws = wb[sht_name]
                            # 遍历所有单元格，确保公式格式正确
                            for row in ws.iter_rows():
                                for cell in row:
                                    if cell.data_type == 'f':  # 'f' 表示公式
                                        formula = cell.value
                                        # 确保公式以 = 开头
                                        if formula and not str(formula).startswith('='):
                                            cell.value = f"={formula}"
                
                # 设置工作簿属性，强制Excel在打开时重新计算所有公式
                wb.calculation.calcMode = 'auto'
                wb.calculation.calcCompleted = False
                wb.calculation.fullCalcOnLoad = True
        
        # 保存文件
        wb.save(file_path)
